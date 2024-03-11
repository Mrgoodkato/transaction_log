import os
import df_processing as dfp
import openpyxl as op
from openpyxl.formatting import Rule
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule



entries = dfp.entryParseXL(os.listdir('files'))
db = dfp.databaseCreation(entries)

database = {}

for entry in db:

    db_entry = db[entry]
    db_entry['Errors'] = ''

    tally_out_list = db_entry.loc[:, 'Tally Out'].unique()

    tally_out_list

    for tout in tally_out_list:

        database[tout] = dfp.processDatabase(db_entry.loc[db_entry['Tally Out'] == tout])

        for error in database[tout].loc[:, ['Indexes', 'Error']].iterrows():

            if error[1]['Error'] == True:

                db_entry.loc[error[1]['Indexes'], 'Error'] = True

        dfp.saveFinalAnalysis(database[tout], entry, tout)
    
    file_path = 'output/{}_log_interpretaion.xlsx'.format(entry)
    db_entry.to_excel(file_path, index=False)

    workbook = op.load_workbook(file_path)
    sheet = workbook.active
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 10
    sheet.column_dimensions['H'].width = 10
    sheet.column_dimensions['I'].width = 10
    
    action_col_style_tallyout = {
        'bg': PatternFill(bgColor="FF006600"),
        'font': Font(color='FFEEECE1')
    }

    action_col_style_rollback = {
        'bg': PatternFill(bgColor="FFFF3300"),
        'font': Font(color='FFEEECE1')
    }

    qty_col_style = {
        'bg': PatternFill(bgColor="FFDA9694"),
        'font': Font(color='FFC00000')
    }

   

    workbook.save(file_path)
