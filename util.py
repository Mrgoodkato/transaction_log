import openpyxl as op
from openpyxl.styles import Font, PatternFill

#Function used to modify the storage object, provide initial values
def modifyStorageObject(data):

    return {
            
            'Action': data['tallyout'].iloc[0],
            'Indexes': [data['key']],
            'Item Nos': [data['tallyout'].iloc[1]],
            'Tally Ins': [data['tallyout'].iloc[2]],
            'Qty': data['qty'],
            'Error': data['err'],
            'Err-location': data['err-location']

        }

#Function that creates a unique value list from a list that might have duplicate values
def uniqueList(list):

    unique_list = []

    for itm in list:
        if itm not in unique_list:
            unique_list.append(itm)
    
    return unique_list

#Function returns True if there are duplicates in the list, False if there are no duplicates
def duplicateFinder(list):

    unique_list = []

    for itm in list:
        if itm not in unique_list:
            unique_list.append(itm)
        else:
            return True
        
    return False

#Function that removes the necessary rows and columns from the Transaction log report
def setupExcel(path):

    wb = op.load_workbook(path)
    ws = wb.active

    ws.unmerge_cells('B2:D2')
    ws.unmerge_cells('B4:J4')
    
    num = 6
    while True:
        try:
            ws.unmerge_cells(f'D{num}:E{num}')
            ws.unmerge_cells(f'J{num}:K{num}')
            num += 1
        except:
            break

    ws.delete_rows(1, 5)
    ws.delete_cols(1, 1)
    ws.delete_cols(4, 1)
    ws.delete_cols(9, 1)

    wb.save('temp/modified.xlsx')

#Helper function that changes the format of cells in the final excel file to highlight important information based on a rule and style
def cellFormatChanger(format, col, rule, ws):

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):

        for cell in row:
            if(cell.value == rule):
                cell.font = format['font']
                cell.fill = format['bg']

#Function that styles the entire worksheet being worked for readability
def excelStyler(ws):

    action_col_style_tallyout = {
    'bg': PatternFill(start_color="FF006600", end_color="FF006600", fill_type="solid"),
    'font': Font(color='FFEEECE1')
    }

    action_col_style_rollback = {
        'bg': PatternFill(start_color="FFFF3300", end_color="FFFF3300", fill_type="solid"),
        'font': Font(color='FFEEECE1')
    }

    qty_col_style = {
        'bg': PatternFill(start_color="FFDA9694", end_color="FFDA9694", fill_type="solid"),
        'font': Font(color='FFC00000')
    }

    error_style = {
        'bg': PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
        'font': Font(color='FFFFFFFF')
    }


    cellFormatChanger(action_col_style_tallyout, 1, 'Tally out movement', ws)
    cellFormatChanger(action_col_style_rollback, 1, 'Rollback movement', ws)
    cellFormatChanger(qty_col_style, 5, 0, ws)
    cellFormatChanger(qty_col_style, 6, 0, ws)
    cellFormatChanger(qty_col_style, 7, 0, ws)
    cellFormatChanger(qty_col_style, 8, 0, ws)
    cellFormatChanger(error_style, 11, True, ws)